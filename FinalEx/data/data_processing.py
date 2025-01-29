import pandas as pd
from openpyxl import load_workbook

def remove_columns_with_values_in_row_and_add_sheet(filepath, sheet_name="Sheet1", target_row=3, new_sheet_name="Processed_Data"):
    """
    指定した行に値が含まれる列を削除し、新しいシートに加工後のデータを書き込む。

    Parameters:
        filepath (str): 読み込みたいExcelファイルのパス
        sheet_name (int/str, optional): 読み込むシート名またはインデックス（デフォルトは最初のシート）
        target_row (int): 値をチェックする行番号（1から始まる）
        new_sheet_name (str): 新しいシート名（デフォルトは "Processed_Data"）
    """
    # Excelデータを読み込む（ヘッダーなしで全ての行を取得）
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)

    # データの行数を確認
    if target_row > len(df):
        raise ValueError(f"指定された行 (target_row={target_row}) はデータの範囲を超えています。")

    # ターゲット行を取得（0インデックスに合わせるため -1）
    row_values = df.iloc[target_row - 1]

    # ターゲット行に値が含まれる列を特定
    columns_to_remove = row_values[row_values.notna()].index.tolist()  # 値がNaNでない列インデックスをリストで取得

    # 特定した列を削除
    processed_df = df.drop(columns=columns_to_remove)

    # 元のExcelファイルを読み込む
    workbook = load_workbook(filepath)

    # 新しいシートを追加
    if new_sheet_name in workbook.sheetnames:
        print(f"シート名 '{new_sheet_name}' は既に存在します。別の名前を指定してください。")
        return

    worksheet = workbook.create_sheet(title=new_sheet_name)

    # 新しいシートに加工後のデータを書き込む
    for row_idx, row_data in enumerate(processed_df.itertuples(index=False, name=None), start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # 上書き保存
    workbook.save(filepath)

    print(f"削除した列: {columns_to_remove}")
    print(f"加工後のデータを新しいシート '{new_sheet_name}' に書き込みました。")

# 使用例
filepath = "data.xlsx"  # 実際のExcelファイルパスを指定
target_row = 3  # 3行目を基準にする
new_sheet_name = "Processed_Data"  # 新しいシート名

# 処理を実行
remove_columns_with_values_in_row_and_add_sheet(filepath, target_row=target_row, new_sheet_name=new_sheet_name)
